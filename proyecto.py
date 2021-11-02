#Nombre: Evan Jesus Tejada Duarte
#Carné: 0907-21-7854

#importo el documento inventario.xlsx
#from _typeshed import Self
from re import S
import openpyxl
from openpyxl import load_workbook
libro = openpyxl.load_workbook("inventario.xlsx")
hoja_productos = libro['Productos']
hoja_clientes = libro['Clientes']
hoja_pedidos = libro['Pedidos']



def pricipal():
    print("1. Productos")
    print("2. Clientes")
    print("3. Pedidos")
    print("4. Informes")
    print("5. Varios")
    print("6. Cancelar \n")

pricipal()
opcionElegida = int(input('Eliga una opción: '))

while opcionElegida != 6:

#1. Productos
    if opcionElegida == 1:
        def produc():
            print('Seleccione la opcion deseada')
            print("1.1 Agregar producto")
            print("1.2 Editar producto")
            print("1.3 Eliminar producto")
            print("1.4 Listar productos")
            print("1.5 Enviar cotización por correo")
            print("1.6 Cancelar \n")
        produc()
        productos =  float(input('Eliga que acción realizar: '))
    #aqui comienza las funciones de las opciones 1.#

    #1.1 Agregar producto
        if productos == 1.1:
            hoja_productos['A1'].value = 'nombre'
            hoja_productos['B1'].value = 'precio'
            hoja_productos['C1'].value = 'existencia'
            addprod = []
            addproddic = {}
            
            producaddpro = str(input("Introduce el nombre del producto: "))
            addproddic['nombre'] = producaddpro
            priceaddpro = float(input("Introduce el precio del producto: "))
            addproddic['precio'] = priceaddpro
            amountaddpro = int(input("Introduce la existencia del producto: "))
            addproddic['existencia'] = amountaddpro
            addprod.append(addproddic)

            firstrow = hoja_productos.max_row + 1
            for addprodin in addprod:
                hoja_productos['A' + str(firstrow)].value = addprodin['nombre']
                hoja_productos['B' + str(firstrow)].value = addprodin['precio']
                hoja_productos['C' + str(firstrow)].value = addprodin['existencia']
            # libro.remove('inventario.xlsx') 
            # libro.save("inventario.xlsx")
    # Esto me paso con todos, no logro hacer que vuelva a las opciones #.# asi que lo mando al menu principal
            opcionElegida = 0
            print("saliendo \n")
            pricipal()
            print("Eliga una opción")
            opcionElegida = int(input())
            
    #1.2 Editar producto
        elif productos == 1.2:
            editprod = []
            editproddic = {}

            for rows in range(2, hoja_productos.max_row + 1):
                producaddpro = hoja_productos['A' + str(rows)].value
                priceaddpro = hoja_productos['B' + str(rows)].value
                amountaddpro = hoja_productos['C' + str(rows)].value
                editproddic['producto'] = priceaddpro
                editproddic['precio'] = priceaddpro
                editproddic['existencia'] = amountaddpro
                editprod.append(editproddic)
                editproddic = {}
            integratoredit = 1
            for editprodin in editprod:
                print('line: '+str(integratoredit),
                ' Producto: '+str(editprodin['producto']),
                ' Precio: '+str(editprodin['precio']),
                ' Existencia: '+str(editprodin['existencia'])
                )
                integratoredit += 1
            editselect = int(input('Ingrese el numero de liena del producto a editar: '))
            edit = editselect + 1
            hoja_productos['A1'].value = 'Producto'
            hoja_productos['B1'].value = 'Precio'
            hoja_productos['C1'].value = 'Existencia'
            editprodadd = []
            editproddicadd = {}
            produceditadd = input('Ingrese el nuevo nombre del producto: '+hoja_productos['A'+str(edit)].value)
            editproddicadd['producto'] = produceditadd
            priceeditadd = float(input('Ingrese el nuevo precio del producto: '+str(hoja_productos['B'+str(edit)].value)))
            editproddicadd['precio'] = priceeditadd
            amounteditadd = float(input('Ingrese el nuevo precio del producto: '+str(hoja_productos['C'+str(edit)].value)))
            editproddicadd['existencia'] = amounteditadd
            editprodadd.append(editproddicadd)
            for editprodaddin in editproddicadd:
                hoja_productos['A'+str(edit)].value = editprodaddin['producto'] 
                hoja_productos['B'+str(edit)].value = editprodaddin['precio']
                hoja_productos['C'+str(edit)].value = editprodaddin['existencia']
            libro.remove('inventario.xlsx')
            libro.save('inventario.xlsx')
            opcionElegida = 0
            produc()
            print("Eliga que acción realizar")
            productos =  float(input())

    #1.3 Eliminar producto
        elif productos == 1.3:
            deleteprod = []
            deleteproddic = {}
            for row in range(2, hoja_productos.max_row + 1):
                proddelete = hoja_productos['A'+str(row)].value
                pricedelete = hoja_productos['B'+str(row)].value
                amountdelete = hoja_productos['C'+str(row)].value
                deleteproddic['producto'] = proddelete
                deleteproddic['precio'] = pricedelete
                deleteproddic['existencia'] = amountdelete
                deleteprod.append(deleteproddic)
                deleteproddic = {}
            integratordelete = 1
            for deleteprodin in deleteprod:
                print('Linea: '+str(deleteprodin),'Prodcuto: '+deleteprodin['producto'],'Precio: '+str(deleteprodin['precio']),'Existencia: '+str(deleteprodin['existencia']))
                integratordelete = integratordelete + 1
            rowdelete = int(input('Ingrese el numero de la fila a borrar: '))
            deleterow = deleteprod + 1
            hoja_productos.delete_rows(deleterow, 1)
            libro.remove('inventario.xlsx')
            libro.save('inventario.xlsx')

            opcionElegida = 0
            print("saliendo \n")
            pricipal()
            print("Eliga una opción")
            opcionElegida = int(input())

    #1.4 Listar productos
        elif productos == 1.4:
            listprod= []
            listproddic = {}
            for row in range(2, hoja_productos.max_row + 1):
                prodlist = hoja_productos["A" + str(row)].value
                pricelist = hoja_productos["B" + str(row)].value
                amountlist = hoja_productos["C" + str(row)].value
                listproddic['nombre'] = prodlist
                listproddic['precio'] = pricelist
                listproddic['existencia'] = amountlist
                listprod.append(listproddic)
                listproddic = {}

            for item in listprod:
                print("Nombre:" + item['nombre'])
                print("Precio:" + str(item['precio']))
                print("Existencia:" + str(item['existencia']))

            opcionElegida = 0
            print("saliendo \n")
            pricipal()
            print("Eliga una opción")
            opcionElegida = int(input())

    #1.5 Enviar cotización por correo
        #cuando haya diccionario remplazar variables
        elif productos == 1.5:
            nombreac = input("Nombre del cliente: ")
            productoel = input("productos elegidos: ")
            correo = input("Correo electronico del cliente: ")
            tuple_cc = {
            "El nombre del cliente es": nombreac,
            "Los productos elegidos son": productoel,
            "El correo es": correo}
            print(tuple_cc)

            opcionElegida = 0
            print("saliendo \n")
            produc()
            print("Eliga una opción")
            opcionElegida = int(input())

    #1.6 Cancelar
        elif productos == 1.6:
            opcionElegida = 0
            print("saliendo \n")
            pricipal()
            print("Eliga una opción")
            opcionElegida = int(input())


# 2. Clientes
    elif opcionElegida == 2:
        def client():
            print("2.1 Agregar Cliente")
            print("2.2 Editar Cliente")
            print("2.3 Eliminar Cliente")
            print("2.4 Listar Clientes")
            print("2.5 Cancelar \n")
        client()
        print("Eliga que acción realizar")
        clientes =  float(input())
    #aqui comienza las funciones de las opciones 2.#

    #2.1 Agregar Cliente
        #cuando haya diccionario remplazar variables
        if clientes == 2.1:
            hoja_clientes['A1'].value = 'Nombre'
            hoja_clientes['B2'].value = 'NIT'
            hoja_clientes['C1'].value = 'Direccion'
            addcliet = []
            addclietdic = {}
            nameclient = input('Ingrese el nombre del cliente: ')
            addclietdic['nombre'] = nameclient
            nitclient = input('Ingrese el nit del cliente: ')
            addclietdic['nit'] = nitclient
            directionclient= input('Ingrese la dirección del cliente: ')
            addclietdic['direccion'] = directionclient
            addcliet.append(addclietdic)

            rowclient = hoja_clientes.max_row + 1
            for addclientin in addcliet:
                hoja_clientes['A'+str(rowclient)].value = addclientin['nombre']
                hoja_clientes['B'+str(rowclient)].value = addclientin['nit']
                hoja_clientes['C'+str(rowclient)].value = addclientin['direccion']
            #libro.remove('inventario.xlsx')
            #libro.save('inventario.xlsx')
            opcionElegida = 0
            print("saliendo \n")
            pricipal()
            print("Eliga una opción")
            opcionElegida = int(input())
    
    #2.2 Editar Cliente
        #cuando haya diccionario remplazar variables
        elif clientes == 2.2:
            editclient= []
            editclientdic = {}
            for editclientin in range(2, hoja_clientes.max_row + 1):
                nameedit = hoja_clientes['A'+str(editclientin)].value
                nitedit = hoja_clientes['B'+str(editclientin)].value
                directionedit = hoja_clientes['C'+str(editclientin)].value
                editclientdic['nombre'] = nameedit
                editclientdic['nit'] = nitedit
                editclientdic['direccion'] = directionedit
                editclient.append(editclientdic)
                editclientdic = {}
            intclientedit = 1
            for clienteditin in editclient:
                print('Linea: '+str(intclientedit),
                ' Nombre: '+clienteditin['nombre'],
                ' NIT: '+str(clienteditin['nit']),
                'Dirección: '+str(clienteditin['direccion']))
                intclientedit += 1
            selectedit = int(input('Ingrese el numero de fila a borrar: '))
            clientedit = selectedit + 1
            hoja_clientes['A1'].value = 'nombre'
            hoja_clientes['B1'].value = 'nit'
            hoja_clientes['C1'].value = 'direccion'
            addclientedit = []
            addclienteditdic = {}
            nameclientadd = input('Ingrese el nuevo nombre del cliente: ' + hoja_clientes['A'+str(clientedit)].value)
            addclienteditdic['nombre'] = nameclientadd
            nitclientadd = input('Ingrese el nuevo NIT del cliente: '+hoja_clientes['B'+str(clientedit)].value)
            addclienteditdic['nit'] = nitclientadd
            directionclientadd = input('Ingrese la nueva dirección del cliente: '+hoja_clientes['C'+str(clientedit)].value)
            addclienteditdic['direccion'] = directionclientadd
            addclientedit.append(addclienteditdic)

            for clientedit in addclientedit:
                hoja_clientes['A'+str(clientedit)].value = editclient['nombre']
                hoja_clientes['B'+str(clientedit)].value = editclient['nit']
                hoja_clientes['C'+str(clientedit)].value = editclient['direccion']
            libro.remove('inventario.xlsx')
            libro.save('inventario.xlsx')

            opcionElegida = 0
            print("saliendo \n")
            pricipal()
            print("Eliga una opción")
            opcionElegida = int(input())

    #2.3 Eliminar Cliente
        #cuando haya diccionario remplazar variables
        elif clientes == 2.3:
            deleteclient= []
            deleteclientdic = {}

            for clientdeletein in range(2, hoja_clientes.max_row + 1):
                clientdelete = hoja_clientes['A'+str(clientdeletein)].value
                nitdelete = hoja_clientes['B'+str(clientdeletein)].value
                directiondelete = hoja_clientes['C'+str(clientdeletein)].value
                deleteclientdic['nombre'] =  clientdelete
                deleteclientdic['nit'] = nitdelete
                deleteclientdic['direccion'] = directiondelete
                deleteclient.append(deleteclientdic)
                deleteclientdic = {}
            integratorclientdelete = 1
            for deleteclientin in deleteclient:
                print('Linea: ',str(deleteclientin)+'Nombre: '+deleteclientin['nombre']+'NIT: ',str(deleteclientin['nit'])+'Dirección: ',str(deleteclientin['direccion']))
                integratorclientdelete = 1
            rowclientdelete = int(input('Ingrese el numero de fijla a borrar: '))
            rowdeleteclient = rowclientdelete + 1
            hoja_clientes.delete_rows(rowdeleteclient, 1)
            # libro.remove('inventario.xlsx')
            # libro.save('inventario.xlsx')

            opcionElegida = 0
            print("saliendo \n")
            pricipal()
            print("Eliga una opción")
            opcionElegida = int(input())
    
    #2.4 Listar Clientes
        #A pesar de ser igual en el comando que el 1.4 este imprime al rededor de 8 caracteres (nombre, nit, dirreccion) bacios
        elif clientes == 2.4:
            listclient = []
            listclientdic = {}
            for row in range(2, hoja_clientes.max_row + 1):
                namelist = hoja_clientes["A" + str(row)].value
                nitlist = hoja_clientes["B" + str(row)].value
                directionlist = hoja_clientes["C" + str(row)].value
                listclientdic['nombre'] = namelist
                listclientdic['nit'] = nitlist
                listclientdic['direccion'] = directionlist
                listclient.append(listclientdic)
                listclientdic = {}
            for listclientin in listclient:
                print(" Nombre: " + str(listclientin['nombre']))
                print(" NIT: " + str(listclientin['nit']))
                print(" Dirección: " + str(listclientin["direccion"]))

            opcionElegida = 0
            print("saliendo \n")
            pricipal()
            print("Eliga una opción")
            opcionElegida = int(input())
    
    #2.5 Cancelar
        if clientes == 2.5:
            opcionElegida = 0
            print("saliendo \n")
            pricipal()
            print("Eliga una opción")
            opcionElegida = int(input())


#3. Pedidos
    elif opcionElegida == 3:
        print("3.1 Agregar Pedido")
        print("3.2 Eliminar Pedido")
        print("3.3 Listar Pedidos")
        print("3.4 Cancelar \n")
        print("Eliga que acción realizar")
        pedidos =  float(input())
    #aqui comienza las funciones de las opciones 3.#

    #3.1 Agregar Pedido
        clientorder = []
        clientorderdic = {}

        for rowclientorder in range(2, hoja_clientes.max_row + 1):
            nameclientorder = hoja_clientes['A'+str(rowclientorder)].value
            clientorderdic['nombre'] = nameclientorder
            clientorder.append(clientorderdic)
            clientorderdic = {}
        print('Los clientes son: ')
        for orderclient in clientorder:
            print('El nombre del cliente es: '+str(orderclient['nombre']))
        hoja_pedidos['A1'].value = 'cliente'
        hoja_pedidos['B1'].value = 'producto'
        hoja_pedidos['C1'].value = 'cantidad'
        hoja_pedidos['D1'].value = 'valorpedido'

        addorder = []
        addorderdic = {}
        nameaddorder = input('Ingrese uno de los nombres que se muestran en la lista anterior: ')
        addorderdic['namecliente'] = nameaddorder

        addprodorder = []
        addprodorderdic = {}
        for rowprodorder in range(2, hoja_productos.max_row + 1):
            orderprod = hoja_productos['A'+str(rowprodorder)].value
            addclienteditdic['producto'] = orderprod
            addprodorder.append(addclienteditdic)
            addclienteditdic = {}

        integratororder = 1
        for prodorder in addorder:
            print('Numero de linea: '+str(integratororder)+'Producto: '+prodorder['producto'])
            integratororder += 1
        nameorderprod = int(input('Ingrese e numero de la linea del prodcuto: '))
        rownameorderprod = nameorderprod + 1
        addorderdic['nameproducto'] = hoja_productos['A'+str(rownameorderprod)].value
        amountorder = int(input('Ingrese la cantidad a comprar: '))
        addorderdic['cantidad'] = amountorder
        addorderdic['precio'] = amountorder * hoja_productos['B'+str(rownameorderprod)].value
        addorder.append(addorderdic)

        rowaddorder = hoja_pedidos.max_row + 1
        for orderin in addorder:
            hoja_pedidos['A'+str(rowaddorder)].value = orderin['namecliente']
            hoja_pedidos['B'+str(rowaddorder)].value = orderin['nameprodcuto']
            hoja_pedidos['C'+str(rowaddorder)].value = orderin['cantidad']
            hoja_pedidos['D'+str(rowaddorder)].value = orderin['precio']
        editamount = hoja_productos['C'+str(rownameorderprod)].value - amountorder
        hoja_productos['C'+str(rownameorderprod)] = editamount
        libro.remove('inventario.xlsx')
        libro.save('invenatrio.xlsx')
        


        opcionElegida = 0
        print("saliendo \n")
        pricipal()
        print("Eliga una opción")
        opcionElegida = int(input())

    #3.2 Eliminar Pedido
        if pedidos == 3.2:
            deleteorder = []
            deleteorderdic = {}

            for rowdeleteorder in range(2, hoja_pedidos.max_row + 1):
                nameclientdeleteorder = hoja_pedidos['A'+str(rowdeleteorder)].value
                nameproddeleteorder = hoja_pedidos['B'+str(rowdeleteorder)].value
                amountdeleteorder = hoja_pedidos['C'+str(rowdeleteorder)].value
                pricedeleteorder = hoja_pedidos['D'+str(rowdeleteorder)].value
                deleteorderdic['namecliente'] = nameclientdeleteorder
                deleteorderdic['nameprodcuto'] = nameproddeleteorder
                deleteorderdic['cantidad'] = amountdeleteorder
                deleteorderdic['precio'] = pricedeleteorder
                deleteorder.append(deleteclientdic)
                deleteclientdic = {}
            integratordeleteorder = 1
            
            for deleteorderin in deleteorder:
                print('Linea: '+str(integratordeleteorder)+'Cliente: '+deleteclientin['namecliente']+'Producto: '+deleteclientin['nameproducto']+'Cantidad: '+str(deleteclientin['cantidad'])+'Precio del pedido: '+str(deleteclientin['precio']))
                integratordeleteorder += 1 
            deleteorderrow = int(input('Ingrese el numero de la fila para borrar: '))
            deleteroworder = deleteorderrow + 1
            hoja_pedidos.delete_rows(deleteroworder, 1)
            libro.remove('inventario.xlsx')
            libro.save('inventario.xlsx')

            opcionElegida = 0
            print("saliendo \n")
            pricipal()
            print("Eliga una opción")
            opcionElegida = int(input())
    
    #3.3 Listar Pedidos
        #esta lista no pide el contenido del .xlsx pero si cambio 'hoja_pedidos' por 'hoja_productos' si es funcional, no lo comprendo
        elif pedidos == 3.3:
            listorder = []
            listorderdic = {}

            for row in range(2, hoja_pedidos.max_row + 1):
                namelistorderclient = hoja_pedidos["A" + str(row)].value
                namelistorderprod = hoja_pedidos["B" + str(row)].value
                amountlistorder = hoja_pedidos["C" + str(row)].value
                pricelistorder = hoja_pedidos["D" + str(row)].value
                listorderdic['namecliente'] =  namelistorderclient
                listorderdic['nameproducto'] = namelistorderprod
                listorderdic['cantidad'] = amountlistorder
                listorderdic['precio'] = pricelistorder
                listorder.append(listorderdic)
                listorderdic = {}

            for listorderin in listorder:
                print("Nombre del Cliente:" + listorderin[' namecliente'])
                print("Nombre del Producto:" + listorderin['nameproducto'])
                print("Cantidad de Producto:" + str(listorderin['cantidad']))
                print("Valor de Prodcuto:" + str(listorderin['valor']))

            opcionElegida = 0
            print("saliendo \n")
            pricipal()
            print("Eliga una opción")
            opcionElegida = int(input())
    
    #3.4 Cancelar
        elif pedidos == 3.4:
            opcionElegida = 0
            print("saliendo \n")
            pricipal()
            print("Eliga una opción")
            opcionElegida = int(input())


#4. Informes
    elif opcionElegida == 4:
        print("4.1 Total de Ventas por Cliente")
        print("4.2 Total de Ventas por Producto")
        print("4.3. Cancelar \n")
        print("Eliga que acción realizar")
        informes =  float(input())
    #aqui comienza las funciones de las opciones 4.#
    #4.1 Total de Ventas por Cliente
        if informes == 4.1:
            clientreport = []
            clientreportdic = {}
            integratorclientreport = 2
            salesclient = 0.0
            for rowclientreport in range(2, hoja_pedidos.max_row + 1):
                nameclientreport = hoja_pedidos['A'+str(rowclientreport)].value
                priceclientreport = hoja_pedidos['D'+str(rowclientreport)].value
                clientreportdic['nameclientreport'] = nameclientreport
                clientreportdic['priceclientreport'] = priceclientreport
                clientreport.append(clientreportdic)
                clientreportdic = {}
            
            for clientreportin in clientreport:
                print('Nombre del cliente: '+clientdeletein['nameclientreport'])
                print('Total de venta: '+str(clientdeletein['priceclientreport']))
                integratorclientreport += 1
            print(salesclient)

            opcionElegida = 0
            print("saliendo \n")
            pricipal()
            print("Eliga una opción")
            opcionElegida = int(input())

    #4.2 Total de Ventas por Producto
        if informes == 4.2:
            prodreport = []
            prodreportdic ={}
            integratorprodreport = 2
            salesprod = 0.0
            for rowprodreport in range(2, hoja_pedidos.max_row + 1):
                nameprodreport = hoja_pedidos['B'+str(rowprodreport)].value
                priceprodreport = hoja_pedidos['D'+str(rowprodreport)].value
                prodreportdic['nameprodreport'] = nameprodreport
                prodreportdic['priceclientreport'] = priceclientreport
                prodreport.append(prodreportdic)
                prodreportdic = {}
            
            for prodreportin in prodreport:
                print('Nombre del producto: '+prodreportin['nameprodreport'])
                print('Total de venta: '+str(prodreportin['priceclientreport']))
                salesprod += hoja_pedidos['D'+str(integratorclientreport)].value
                integratorclientreport +=1

            opcionElegida = 0
            print("saliendo \n")
            pricipal()
            print("Eliga una opción")
            opcionElegida = int(input())

    #4.3 Cancelar
        elif informes == 4.3:
            opcionElegida = 0
            print("saliendo \n")
            pricipal()
            print("Eliga una opción")
            opcionElegida = int(input())


#5. Varios
    elif opcionElegida == 5:
        print("5.1 Crear Copia de Seguridad de Datos")
        print("5.2. Cancelar \n")
        print("Eliga que acción realizar")
        varios =  float(input())
    #aqui comienza las funciones de las opciones 5.#
    #5.1 Crear Copia de Seguridad de Datos
        if varios == 5.1:
            print("Creando copia de seguridad")
            #conciste en un correo a mi correo .umg
            
        opcionElegida = 0
        print("saliendo \n")
        pricipal()
        print("Eliga una opción")
        opcionElegida = int(input())
    
    #5.2 Cancelar
        if varios == 5.2:
            opcionElegida = 0
            print("saliendo \n")
            pricipal()
            print("Eliga una opción")
            opcionElegida = int(input())

#no hay un numero 6 ya que es el que sale del codigo

